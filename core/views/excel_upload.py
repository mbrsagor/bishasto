from openpyxl import load_workbook
from rest_framework import views, status
from rest_framework.response import Response

from core.models.location import Location


class LocationExcelUploadView(views.APIView):
    def post(self, request):
        excel_file = request.data.get('excell_file')
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, location_type, is_active = row
            location = Location.objects.create(name=name, location_type=location_type, is_active=is_active)
            if location is not None:
                location.save()
                return Response('Successfully uploaded', status=status.HTTP_201_CREATED)
            return Response('Upload error', status=status.HTTP_400_BAD_REQUEST)
