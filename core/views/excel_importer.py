from openpyxl import load_workbook

def import_from_excel_view(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            distributor, website, street, email, phone, city, state, zip_code, is_gaseous, is_diesel = row
            Distributor.objects.create(distributor=distributor, website=website, street=street, email=email,
                                       phone=phone, city=city, state=state, zip_code=zip_code, is_gaseous=is_gaseous,
                                       is_diesel=is_diesel)
